# Databricks notebook source
# MAGIC
# MAGIC %run "/Workspace/Repos/deepanshuk36@gmail.com/deltalakerepo/(Clone) deltalake/importfunctions/impfun"

# COMMAND ----------

def getlatest(location):
    a=[]
    
    for path in dbutils.fs.ls(location):
        a.append((path[0],path[0].split('/')[-1][13:].split('.')[0].replace('_','-')))
    a.sort()
    if(len(a))==0:
        return None

    return a[-1]


filename=getlatest('/mnt/sourceeadlsgen2/constructor')[0]
filedate=getlatest('/mnt/sourceeadlsgen2/constructor')[1]

print(filename,filedate)

# COMMAND ----------

pip install openpyxl

# COMMAND ----------

try : import openpyxl
except ImportError:
    %pip install openpyxl
    import openpyxl    





# COMMAND ----------

import openpyxl as xl

import pandas as pd

# path=latestpath('/mnt/sourceblob/')
path=filename

path.replace('dbfs:','/dbfs')


print(path)

sheet_excel=xl.load_workbook(path.replace('dbfs:','/dbfs'))

print(sheet_excel.sheetnames)

from pyspark.sql.functions import lit

df=spark.createDataFrame([],schema="constructorId int,constructorRef string,name string,nationality string,url string,source_location string,load_time timestamp,file_date string")


for sheet in sheet_excel.sheetnames:
    data=pd.read_excel(path.replace('dbfs:','/dbfs'),engine='openpyxl',sheet_name=f'{sheet}')
    tempdf=spark.createDataFrame(data,schema="constructorId int,constructorRef string,name string,	nationality string,url string")
    tempdf=addload_timestamp(tempdf,path)
    tempdf=tempdf.withColumn('file_date',lit(filedate))
    print(tempdf.count())
    df=df.union(tempdf)


if(filedate is not None and filedate in getlatest('/mnt/bronzeadls/constructor/')[0]):
    print(filedate,df["file_date"])
    df.write.partitionBy('file_date').mode('overwrite')\
    .option('replaceWhere', f"file_date=='{filedate}'")\
    .save('/mnt/bronzeadls/constructor/')

else:
    df.write.partitionBy('file_date').mode('append').save('/mnt/bronzeadls/constructor/')




